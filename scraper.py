from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time

# ============================================================
#   GOOGLE MAPS SCRAPER v2 — Fixed!
#   Built by Bibhuti Adhikari — bibhutiportfolio.vercel.app
# ============================================================

SEARCH_QUERY = "restaurants in Kathmandu"
MAX_RESULTS  = 50

# ============================================================

def scrape_google_maps(search_query, max_results):
    print(f"\n{'='*55}")
    print(f"  🗺️  Google Maps Scraper v2")
    print(f"  🔎 Query: {search_query}")
    print(f"  📦 Max  : {max_results}")
    print(f"{'='*55}\n")

    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page    = context.new_page()

        print("  ⏳ Opening Google Maps...")
        page.goto("https://www.google.com/maps", timeout=60000)
        page.wait_for_timeout(3000)

        # Search
        print(f"  🔍 Searching: {search_query}")
        page.fill("input[name='q']", search_query)
        page.keyboard.press("Enter")
        page.wait_for_timeout(5000)

        # Scroll to load more
        print("  📜 Loading results...")
        try:
            feed = page.locator("div[role='feed']")
            for i in range(15):
                feed.evaluate("el => el.scrollTop += 800")
                page.wait_for_timeout(800)
                print(f"  ↓ Scrolling... ({i+1}/15)")
        except:
            pass

        page.wait_for_timeout(2000)

        # Get all hrefs first — before clicking anything
        print("\n  📋 Collecting listing URLs...")
        all_links = page.eval_on_selector_all(
            "a[href*='/maps/place/']",
            "els => els.map(el => el.href)"
        )
        # Remove duplicates
        seen  = set()
        links = []
        for l in all_links:
            if l not in seen:
                seen.add(l)
                links.append(l)

        print(f"  ✅ Found {len(links)} unique listings\n")

        # Visit each URL directly — no stale element issues!
        for idx, url in enumerate(links[:max_results], 1):
            try:
                page.goto(url, timeout=30000)
                page.wait_for_timeout(2500)

                data = {"No.": idx}

                # Name
                try:
                    data["Business Name"] = page.locator("h1").first.inner_text(timeout=3000)
                except:
                    data["Business Name"] = "N/A"

                # Rating
                try:
                    data["Rating"] = page.locator("div.F7nice span[aria-hidden='true']").first.inner_text(timeout=2000)
                except:
                    data["Rating"] = "N/A"

                # Reviews
                try:
                    rev = page.locator("div.F7nice span[aria-label*='review']").first.inner_text(timeout=2000)
                    data["Reviews"] = rev.strip("()")
                except:
                    data["Reviews"] = "N/A"

                # Category
                try:
                    data["Category"] = page.locator("button.DkEaL").first.inner_text(timeout=2000)
                except:
                    try:
                        data["Category"] = page.locator("span.mgr77e").first.inner_text(timeout=2000)
                    except:
                        data["Category"] = "N/A"

                # Address
                try:
                    data["Address"] = page.locator("button[data-item-id='address']").inner_text(timeout=2000)
                except:
                    data["Address"] = "N/A"

                # Phone
                try:
                    data["Phone"] = page.locator("button[data-item-id*='phone']").inner_text(timeout=2000)
                except:
                    data["Phone"] = "N/A"

                # Website
                try:
                    data["Website"] = page.locator("a[data-item-id='authority']").get_attribute("href", timeout=2000)
                except:
                    data["Website"] = "N/A"

                # Open now
                try:
                    data["Open Now"] = page.locator("span.ZDu9vd span").first.inner_text(timeout=2000)
                except:
                    data["Open Now"] = "N/A"

                results.append(data)
                print(f"  [{idx:03d}] {data['Business Name'][:40]:<40} | ⭐{data['Rating']} | {data['Phone']}")

            except Exception as e:
                print(f"  [{idx:03d}] ⚠️  Error: {str(e)[:50]}")
                continue

        browser.close()

    print(f"\n  ✅ Total scraped: {len(results)} businesses")
    return results


def export_excel(data, search_query):
    if not data:
        print("\n  ⚠️  No data!")
        return

    safe     = search_query.replace(" ","_")[:30]
    filename = f"gmaps_{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Google Maps Data"

    for i, w in enumerate([5,35,8,10,20,40,20,35,15], 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    with_phone   = sum(1 for b in data if b.get("Phone","N/A")   != "N/A")
    with_website = sum(1 for b in data if b.get("Website","N/A") != "N/A")

    # Title
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = f"🗺️ {search_query.title()} — {datetime.now().strftime('%B %d, %Y')}"
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1A1A2E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells("A2:I2")
    s = ws["A2"]
    s.value     = f"Total: {len(data)} | Phones: {with_phone} | Websites: {with_website} | Source: Google Maps | By: Bibhuti Adhikari"
    s.font      = Font(bold=True, size=10, color="FFFFFF")
    s.fill      = PatternFill("solid", fgColor="16213E")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["#","Business Name","Rating","Reviews","Category","Address","Phone","Website","Open Now"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor="E55934")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, biz in enumerate(data, 4):
        fg   = "FFF5F2" if i % 2 == 0 else "FFFFFF"
        vals = [biz.get("No.",""), biz.get("Business Name",""),
                biz.get("Rating",""), biz.get("Reviews",""),
                biz.get("Category",""), biz.get("Address",""),
                biz.get("Phone",""), biz.get("Website",""),
                biz.get("Open Now","")]
        for col, val in enumerate(vals, 1):
            c           = ws.cell(i, col, val)
            c.fill      = PatternFill("solid", fgColor=fg)
            c.border    = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1: c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 8: c.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[i].height = 16

    # Summary
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30
    ws2.merge_cells("A1:B1")
    h = ws2["A1"]
    h.value     = "📊 Scrape Summary"
    h.font      = Font(bold=True, size=13, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor="1A1A2E")
    h.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28

    rated = [b for b in data if b.get("Rating","N/A") not in ("N/A","")]
    try:
        avg_r = round(sum(float(b["Rating"]) for b in rated) / len(rated), 2)
    except:
        avg_r = "N/A"

    for row, (k, v) in enumerate([
        ("Total Businesses", len(data)),
        ("With Phone",       with_phone),
        ("With Website",     with_website),
        ("Avg Rating",       avg_r),
        ("Search Query",     search_query),
        ("Source",           "Google Maps"),
        ("Scraped On",       datetime.now().strftime("%B %d, %Y")),
        ("Built by",         "Bibhuti Adhikari"),
        ("Portfolio",        "bibhutiportfolio.vercel.app"),
    ], 2):
        fg = "FFF5F2" if row % 2 == 0 else "FFFFFF"
        ws2[f"A{row}"] = k
        ws2[f"B{row}"] = str(v)
        ws2[f"A{row}"].font = Font(bold=True)
        ws2[f"A{row}"].fill = PatternFill("solid", fgColor=fg)
        ws2[f"B{row}"].fill = PatternFill("solid", fgColor=fg)

    wb.save(filename)
    print(f"\n  ✅ Excel saved → {filename}")
    print(f"  📊 {len(data)} businesses | {with_phone} phones | {with_website} websites")
    return filename


if __name__ == "__main__":
    print("🚀 Google Maps Scraper v2 — Playwright")
    print("   Built by Bibhuti Adhikari")
    print("   bibhutiportfolio.vercel.app\n")

    data = scrape_google_maps(SEARCH_QUERY, MAX_RESULTS)

    if data:
        export_excel(data, SEARCH_QUERY)
        print("\n🎉 Done! Open Excel to see your data.")
    else:
        print("\n❌ No data scraped.")